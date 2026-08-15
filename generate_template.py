import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ============= 样式定义 =============
title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sub_header_font = Font(name='微软雅黑', size=10, bold=True, color='1F3864')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
total_font = Font(name='微软雅黑', size=10, bold=True, color='1F3864')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
profit_positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
profit_negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
note_font = Font(name='微软雅黑', size=9, italic=True, color='808080')
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

THB_FMT = '#,##0.00" THB"'
CNY_FMT = '#,##0.00" CNY"'
PCT_FMT = '0.00%'
INT_FMT = '#,##0'

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell

def merge_title(ws, row, start_col, end_col, text):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    set_cell(ws, row, start_col, text, title_font, title_fill, center_align, thin_border)

def merge_section(ws, row, start_col, end_col, text):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    set_cell(ws, row, start_col, text, sub_header_font, sub_header_fill, left_align, thin_border)

# ==========================================
# Sheet 1: 使用说明
# ==========================================
ws0 = wb.active
ws0.title = '使用说明'
ws0.column_dimensions['A'].width = 3
ws0.column_dimensions['B'].width = 80

rows_info = [
    ('', '📊 TikTok Shop 经营利润核算表 V1.0', True, 'title'),
    ('', '', False, ''),
    ('', '【适用范围】', True, 'sub'),
    ('', 'TikTok Shop 泰国站点经营利润核算，支持多店铺、多SKU、月度/阶段性分析', False, 'normal'),
    ('', '', False, ''),
    ('', '【表格结构】', True, 'sub'),
    ('', '01-经营总览     : 公司整体核心指标看板（填入数据后自动计算）', False, 'normal'),
    ('', '02-利润明细     : 按SKU维度的完整利润明细表（主表）', False, 'normal'),
    ('', '03-店铺汇总     : 按店铺维度的利润对比', False, 'normal'),
    ('', '04-成本结构     : 成本费用拆解与占比分析', False, 'normal'),
    ('', '05-推广分析     : 广告投放效果与利润关联分析', False, 'normal'),
    ('', '06-达人分析     : 达人渠道销售与利润分析', False, 'normal'),
    ('', '07-售后分析     : 退款退货与售后损失分析', False, 'normal'),
    ('', '08-指标定义     : 所有指标的计算公式与口径说明', False, 'normal'),
    ('', '', False, ''),
    ('', '【使用方法】', True, 'sub'),
    ('', '1. 基础数据填写：在"02-利润明细"表中按SKU行填入销售、成本、费用等原始数据', False, 'normal'),
    ('', '2. 黄色背景单元格为需要手动填写的输入项', False, 'normal'),
    ('', '3. 其他所有表格数据自动从明细表汇总计算', False, 'normal'),
    ('', '4. 绿色数字=盈利/正向，红色数字=亏损/负向', False, 'normal'),
    ('', '', False, ''),
    ('', '【核心口径】', True, 'sub'),
    ('', '• 收入口径：GMV → 有效销售额 → 净销售收入 → 平台结算收入（分层核算，不混用）', False, 'normal'),
    ('', '• 利润口径：经营贡献利润 = 净销售收入 - 所有变动成本（不含固定成本）', False, 'normal'),
    ('', '• 固定成本：人工/房租/软件费等暂不计入经营贡献利润，预留位置待确认', False, 'normal'),
    ('', '• 币种：默认泰铢(THB)，可按汇率换算为人民币(CNY)', False, 'normal'),
    ('', '', False, ''),
    ('', '【数据来源】', True, 'sub'),
    ('', 'A. TikTok销售数据     → 后台订单报表/销售报表', False, 'normal'),
    ('', 'B. TikTok结算数据     → 财务中心/结算报表', False, 'normal'),
    ('', 'C. TikTok售后数据     → 售后/退款报表', False, 'normal'),
    ('', 'D. TikTok Ads数据     → 广告后台/GMV Max报表', False, 'normal'),
    ('', 'E. 达人/Affiliate数据 → 达人广场/联盟报表', False, 'normal'),
    ('', 'F. 商品成本数据       → 采购成本表/SKU成本表', False, 'normal'),
    ('', 'G. 仓储物流成本       → 头程/尾程物流表', False, 'normal'),
    ('', 'H. 其他经营费用       → 费用明细表', False, 'normal'),
    ('', '', False, ''),
    ('', '【注意事项】', True, 'sub'),
    ('', '⚠️ 所有百分比指标分母统一为「有效销售额」，避免口径混乱', False, 'normal'),
    ('', '⚠️ 未结算订单仅供参考，正式利润核算以已结算订单为准', False, 'normal'),
    ('', '⚠️ 汇率按账期内平均汇率或结算日汇率，需统一口径', False, 'normal'),
    ('', '⚠️ 平台佣金/交易费/税费等以TikTok实际扣款为准，不用理论费率推算', False, 'normal'),
]

for i, (col_b, text, is_bold, style) in enumerate(rows_info, 1):
    if style == 'title':
        set_cell(ws0, i, 2, text, Font(name='微软雅黑', size=16, bold=True, color='2F5496'), None, left_align)
    elif style == 'sub':
        set_cell(ws0, i, 2, text, Font(name='微软雅黑', size=11, bold=True, color='2F5496'), None, left_align)
    else:
        set_cell(ws0, i, 2, text, Font(name='微软雅黑', size=10, color='333333'), None, left_align)

# ==========================================
# Sheet 2: 经营总览
# ==========================================
ws1 = wb.create_sheet('01-经营总览')
for col in range(1, 10):
    ws1.column_dimensions[get_column_letter(col)].width = 16
ws1.column_dimensions['A'].width = 28

# 标题
merge_title(ws1, 1, 1, 8, 'TikTok Shop 经营利润总览表')
ws1.row_dimensions[1].height = 32

# 基本信息行
set_cell(ws1, 2, 1, '统计周期:', sub_header_font, sub_header_fill, right_align, thin_border)
merge_title_simple = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
set_cell(ws1, 2, 2, '2026年__月', normal_font, merge_title_simple, center_align, thin_border)
set_cell(ws1, 2, 3, '统计币种:', sub_header_font, sub_header_fill, right_align, thin_border)
set_cell(ws1, 2, 4, 'THB', normal_font, merge_title_simple, center_align, thin_border)
set_cell(ws1, 2, 5, '汇率(THB→CNY):', sub_header_font, sub_header_fill, right_align, thin_border)
set_cell(ws1, 2, 6, '0.200', normal_font, merge_title_simple, center_align, thin_border)
set_cell(ws1, 2, 7, '店铺数量:', sub_header_font, sub_header_fill, right_align, thin_border)
set_cell(ws1, 2, 8, '__家', normal_font, merge_title_simple, center_align, thin_border)

# 核心指标区 - 第一行大指标
row = 4
metrics_main = [
    ('GMV总额', '=SUM(利润明细!G:G)', THB_FMT),
    ('有效销售额', '=SUM(利润明细!H:H)', THB_FMT),
    ('净销售收入', '=SUM(利润明细!K:K)', THB_FMT),
    ('经营贡献利润', '=SUM(利润明细!Z:Z)', THB_FMT),
]
merge_section(ws1, row, 1, 8, '📈 核心经营指标')
row += 1
for i, (name, formula, fmt) in enumerate(metrics_main):
    col = i * 2 + 1
    set_cell(ws1, row, col, name, header_font, header_fill, center_align, thin_border)
    ws1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    set_cell(ws1, row+1, col, formula, Font(name='微软雅黑', size=13, bold=True, color='2F5496'), None, center_align, thin_border)
    ws1.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws1.cell(row=row+1, column=col).number_format = fmt
row += 2

# 第二行指标
metrics_second = [
    ('订单量', '=SUM(利润明细!E:E)', INT_FMT),
    ('销量(件)', '=SUM(利润明细!F:F)', INT_FMT),
    ('客单价', '=IFERROR(SUM(利润明细!H:H)/SUM(利润明细!E:E),0)', THB_FMT),
    ('经营利润率', '=IFERROR(SUM(利润明细!Z:Z)/SUM(利润明细!H:H),0)', PCT_FMT),
]
for i, (name, formula, fmt) in enumerate(metrics_second):
    col = i * 2 + 1
    set_cell(ws1, row, col, name, sub_header_font, sub_header_fill, center_align, thin_border)
    ws1.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    set_cell(ws1, row+1, col, formula, total_font, None, center_align, thin_border)
    ws1.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws1.cell(row=row+1, column=col).number_format = fmt
row += 2

# 收入结构
merge_section(ws1, row, 1, 8, '💰 收入结构')
row += 1
income_items = [
    ('GMV总额', '=SUM(利润明细!G:G)', THB_FMT, '所有订单含退款的成交总额'),
    ('减：退款金额', '=SUM(利润明细!I:I)', THB_FMT, '退货+仅退款+取消'),
    ('有效销售额', '=SUM(利润明细!H:H)', THB_FMT, 'GMV - 退款（核心收入口径）'),
    ('减：卖家折扣/促销', '=SUM(利润明细!J:J)', THB_FMT, '商家承担的折扣和促销费用'),
    ('净销售收入', '=SUM(利润明细!K:K)', THB_FMT, '有效销售额 - 卖家折扣'),
    ('买家支付运费', '=SUM(利润明细!L:L)', THB_FMT, '买家承担的运费部分'),
]
set_cell(ws1, row, 1, '项目', header_font, header_fill, center_align, thin_border)
set_cell(ws1, row, 2, '金额(THB)', header_font, header_fill, center_align, thin_border)
set_cell(ws1, row, 3, '占有效销售额比', header_font, header_fill, center_align, thin_border)
ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
set_cell(ws1, row, 4, '说明', header_font, header_fill, left_align, thin_border)
row += 1
for name, formula, fmt, note in income_items:
    is_total = name in ('有效销售额', '净销售收入')
    f = total_font if is_total else normal_font
    fl = total_fill if is_total else None
    set_cell(ws1, row, 1, name, f, fl, left_align, thin_border)
    set_cell(ws1, row, 2, formula, f, fl, right_align, thin_border)
    ws1.cell(row=row, column=2).number_format = fmt
    ratio_formula = f'=IFERROR(B{row}/SUM(利润明细!H:H),0)'
    set_cell(ws1, row, 3, ratio_formula, f, fl, center_align, thin_border)
    ws1.cell(row=row, column=3).number_format = PCT_FMT
    ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    set_cell(ws1, row, 4, note, note_font, fl, left_align, thin_border)
    row += 1

row += 1
# 成本结构
merge_section(ws1, row, 1, 8, '💸 成本费用结构')
row += 1
cost_items = [
    ('商品成本', '=SUM(利润明细!M:M)', THB_FMT, '采购成本/出厂价'),
    ('平台费用合计', '=SUM(利润明细!P:P)+SUM(利润明细!Q:Q)+SUM(利润明细!R:R)', THB_FMT, '佣金+交易费+税费'),
    ('  平台佣金', '=SUM(利润明细!P:P)', THB_FMT, ''),
    ('  交易手续费', '=SUM(利润明细!Q:Q)', THB_FMT, ''),
    ('  平台税费', '=SUM(利润明细!R:R)', THB_FMT, 'VAT等'),
    ('推广费用合计', '=SUM(利润明细!S:S)', THB_FMT, 'TikTok Ads + GMV Max'),
    ('达人佣金合计', '=SUM(利润明细!T:T)', THB_FMT, ''),
    ('物流费用合计', '=SUM(利润明细!U:U)+SUM(利润明细!V:V)', THB_FMT, '头程+尾程'),
    ('  头程物流', '=SUM(利润明细!U:U)', THB_FMT, '国内到泰国'),
    ('  尾程物流', '=SUM(利润明细!V:V)', THB_FMT, '泰国本地配送'),
    ('仓储费用', '=SUM(利润明细!W:W)', THB_FMT, '海外仓/仓储费'),
    ('售后损失', '=SUM(利润明细!Y:Y)', THB_FMT, '退款中商家承担部分'),
    ('其他经营费用', '=SUM(利润明细!X:X)', THB_FMT, '包材/中转/提现费等'),
]
set_cell(ws1, row, 1, '成本项目', header_font, header_fill, center_align, thin_border)
set_cell(ws1, row, 2, '金额(THB)', header_font, header_fill, center_align, thin_border)
set_cell(ws1, row, 3, '占有效销售额比', header_font, header_fill, center_align, thin_border)
ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
set_cell(ws1, row, 4, '说明', header_font, header_fill, left_align, thin_border)
row += 1
for name, formula, fmt, note in cost_items:
    is_subtotal = '合计' in name
    f = total_font if is_subtotal else normal_font
    fl = total_fill if is_subtotal else None
    indent = '      ' if name.startswith('  ') else ''
    set_cell(ws1, row, 1, indent + name.strip(), f, fl, left_align, thin_border)
    set_cell(ws1, row, 2, formula, f, fl, right_align, thin_border)
    ws1.cell(row=row, column=2).number_format = fmt
    ratio_formula = f'=IFERROR(B{row}/SUM(利润明细!H:H),0)'
    set_cell(ws1, row, 3, ratio_formula, f, fl, center_align, thin_border)
    ws1.cell(row=row, column=3).number_format = PCT_FMT
    ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    set_cell(ws1, row, 4, note, note_font, fl, left_align, thin_border)
    row += 1

# 总变动成本
set_cell(ws1, row, 1, '总变动成本', total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), left_align, thin_border)
total_cost_formula = f'=SUM(利润明细!M:M)+SUM(利润明细!P:P)+SUM(利润明细!Q:Q)+SUM(利润明细!R:R)+SUM(利润明细!S:S)+SUM(利润明细!T:T)+SUM(利润明细!U:U)+SUM(利润明细!V:V)+SUM(利润明细!W:W)+SUM(利润明细!X:X)+SUM(利润明细!Y:Y)'
set_cell(ws1, row, 2, total_cost_formula, total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), right_align, thin_border)
ws1.cell(row=row, column=2).number_format = THB_FMT
set_cell(ws1, row, 3, f'=IFERROR(B{row}/SUM(利润明细!H:H),0)', total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), center_align, thin_border)
ws1.cell(row=row, column=3).number_format = PCT_FMT
ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
set_cell(ws1, row, 4, '以上所有变动成本之和', note_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), left_align, thin_border)
row += 2

# 利润计算
merge_section(ws1, row, 1, 8, '📊 利润计算')
row += 1
profit_items = [
    ('净销售收入', '=SUM(利润明细!K:K)', THB_FMT, ''),
    ('减：商品成本', '=SUM(利润明细!M:M)', THB_FMT, ''),
    ('毛利润', '=SUM(利润明细!K:K)-SUM(利润明细!M:M)', THB_FMT, '净收入 - 商品成本'),
    ('毛利率', '=IFERROR((SUM(利润明细!K:K)-SUM(利润明细!M:M))/SUM(利润明细!H:H),0)', PCT_FMT, '÷ 有效销售额'),
    ('减：所有变动成本（含商品成本）', total_cost_formula, THB_FMT, ''),
    ('经营贡献利润', '=SUM(利润明细!Z:Z)', THB_FMT, '净收入 - 所有变动成本'),
    ('经营贡献利润率', '=IFERROR(SUM(利润明细!Z:Z)/SUM(利润明细!H:H),0)', PCT_FMT, '÷ 有效销售额'),
]
for name, formula, fmt, note in profit_items:
    is_key = name in ('毛利润', '经营贡献利润', '毛利率', '经营贡献利润率')
    f = Font(name='微软雅黑', size=11, bold=True, color='C00000') if is_key and '利润' in name else (total_font if is_key else normal_font)
    fl = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') if is_key else None
    set_cell(ws1, row, 1, name, f, fl, left_align, thin_border)
    set_cell(ws1, row, 2, formula, f, fl, right_align, thin_border)
    ws1.cell(row=row, column=2).number_format = fmt
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    set_cell(ws1, row, 3, note, note_font, fl, left_align, thin_border)
    row += 1

row += 1
# 固定成本预留
merge_section(ws1, row, 1, 8, '🏢 固定成本（预留，暂不计入经营利润）')
row += 1
fixed_items = [
    ('人工成本', '', THB_FMT, '待确认是否计入'),
    ('办公场地/房租', '', THB_FMT, ''),
    ('软件/工具订阅费', '', THB_FMT, ''),
    ('管理费用', '', THB_FMT, ''),
    ('其他固定成本', '', THB_FMT, ''),
]
for name, formula, fmt, note in fixed_items:
    set_cell(ws1, row, 1, name, normal_font, None, left_align, thin_border)
    set_cell(ws1, row, 2, formula if formula else '待确认', note_font, None, right_align, thin_border)
    if formula: ws1.cell(row=row, column=2).number_format = fmt
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    set_cell(ws1, row, 3, note, note_font, None, left_align, thin_border)
    row += 1

# ==========================================
# Sheet 3: 利润明细表（主表）
# ==========================================
ws2 = wb.create_sheet('02-利润明细')

# 标题
merge_title(ws2, 1, 1, 26, 'TikTok Shop 利润明细表（按SKU维度）')
ws2.row_dimensions[1].height = 30

# 表头行
row = 2
headers = [
    # (列号, 名称, 宽度, 类型)
    (1, '序号', 6, 'text'),
    (2, '店铺名称', 16, 'input'),
    (3, '商品ID', 14, 'input'),
    (4, 'SKU编码', 16, 'input'),
    (5, '订单量', 10, 'input'),
    (6, '销量(件)', 10, 'input'),
    (7, 'GMV', 14, 'input'),
    (8, '有效销售额', 14, 'calc'),   # GMV - 退款
    (9, '退款金额', 14, 'input'),
    (10, '卖家折扣/促销', 14, 'input'),
    (11, '净销售收入', 14, 'calc'),  # 有效销售 - 折扣
    (12, '买家付运费', 12, 'input'),
    (13, '商品成本', 14, 'input'),
    (14, '单均商品成本', 12, 'calc'),
    (15, '毛利率', 10, 'calc'),
    (16, '平台佣金', 12, 'input'),
    (17, '交易手续费', 12, 'input'),
    (18, '平台税费(VAT)', 12, 'input'),
    (19, '推广费用', 14, 'input'),
    (20, '达人佣金', 12, 'input'),
    (21, '头程物流', 12, 'input'),
    (22, '尾程物流', 12, 'input'),
    (23, '仓储费用', 12, 'input'),
    (24, '其他费用', 12, 'input'),
    (25, '售后损失', 12, 'input'),
    (26, '经营贡献利润', 16, 'calc'),
    (27, '经营利润率', 12, 'calc'),
    (28, 'ROI', 10, 'calc'),
]

# 分组标记 (列起始, 列结束, 组名)
groups = [
    (1, 4, '基础信息'),
    (5, 7, '销售数据'),
    (8, 12, '收入计算'),
    (13, 15, '商品成本&毛利'),
    (16, 18, '平台费用'),
    (19, 20, '推广&达人'),
    (21, 25, '物流&仓储&其他'),
    (26, 28, '利润指标'),
]

# 分组头
for start, end, name in groups:
    ws2.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    set_cell(ws2, row, start, name, sub_header_font, sub_header_fill, center_align, thin_border)
row += 1

# 列表头
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

for col, name, width, htype in headers:
    ws2.column_dimensions[get_column_letter(col)].width = width
    f = input_fill if htype == 'input' else calc_fill
    set_cell(ws2, row, col, name, header_font, header_fill, center_align, thin_border)

# 冻结
ws2.freeze_panes = 'E4'

# 示例数据行 + 公式
for r in range(4, 103):  # 100行数据
    # 序号
    set_cell(ws2, r, 1, r-3, normal_font, None, center_align, thin_border)
    # 输入列留空（黄色）
    input_cols = [2,3,4,5,6,7,9,10,12,13,16,17,18,19,20,21,22,23,24,25]
    for c in input_cols:
        set_cell(ws2, r, c, '', normal_font, input_fill, center_align if c in [5,6] else left_align, thin_border)
    # 金额格式
    for c in [7,9,10,12,13,16,17,18,19,20,21,22,23,24,25]:
        ws2.cell(row=r, column=c).number_format = THB_FMT
    # 计算列
    # 有效销售额 = GMV - 退款
    set_cell(ws2, r, 8, f'=IFERROR(G{r}-I{r},0)', normal_font, calc_fill, right_align, thin_border)
    ws2.cell(row=r, column=8).number_format = THB_FMT
    # 净销售收入 = 有效销售 - 卖家折扣
    set_cell(ws2, r, 11, f'=IFERROR(H{r}-J{r},0)', normal_font, calc_fill, right_align, thin_border)
    ws2.cell(row=r, column=11).number_format = THB_FMT
    # 单均商品成本
    set_cell(ws2, r, 14, f'=IFERROR(M{r}/F{r},0)', normal_font, calc_fill, right_align, thin_border)
    ws2.cell(row=r, column=14).number_format = THB_FMT
    # 毛利率 = (净收入-商品成本)/有效销售
    set_cell(ws2, r, 15, f'=IFERROR((K{r}-M{r})/H{r},0)', normal_font, calc_fill, center_align, thin_border)
    ws2.cell(row=r, column=15).number_format = PCT_FMT
    # 经营贡献利润 = 净销售收入 - 商品成本 - 平台费 - 推广 - 达人 - 物流 - 仓储 - 其他 - 售后损失
    set_cell(ws2, r, 26, f'=IFERROR(K{r}-M{r}-P{r}-Q{r}-R{r}-S{r}-T{r}-U{r}-V{r}-W{r}-X{r}-Y{r},0)', normal_font, calc_fill, right_align, thin_border)
    ws2.cell(row=r, column=26).number_format = THB_FMT
    # 经营利润率 = 利润/有效销售
    set_cell(ws2, r, 27, f'=IFERROR(Z{r}/H{r},0)', normal_font, calc_fill, center_align, thin_border)
    ws2.cell(row=r, column=27).number_format = PCT_FMT
    # ROI = 有效销售 / (推广费用+商品成本)
    set_cell(ws2, r, 28, f'=IFERROR(H{r}/(M{r}+S{r}),0)', normal_font, calc_fill, center_align, thin_border)
    ws2.cell(row=r, column=28).number_format = '0.00'

# 汇总行
total_row = 103
set_cell(ws2, total_row, 1, '合计', total_font, total_fill, center_align, thin_border)
ws2.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
for c in [5,6,7,9,10,12,13,16,17,18,19,20,21,22,23,24,25]:
    set_cell(ws2, total_row, c, f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}102)', total_font, total_fill, right_align, thin_border)
    ws2.cell(row=total_row, column=c).number_format = THB_FMT if c not in [5,6] else INT_FMT
# 计算列汇总
for c, formula in [
    (8, f'=SUM(H4:H102)'),
    (11, f'=SUM(K4:K102)'),
    (14, f'=IFERROR(M{total_row}/F{total_row},0)'),
    (15, f'=IFERROR((K{total_row}-M{total_row})/H{total_row},0)'),
    (26, f'=SUM(Z4:Z102)'),
    (27, f'=IFERROR(Z{total_row}/H{total_row},0)'),
    (28, f'=IFERROR(H{total_row}/(M{total_row}+S{total_row},0)'),
]:
    set_cell(ws2, total_row, c, formula, total_font, total_fill, right_align if c not in [15,27,28] else center_align, thin_border)
    if c in [8,11,14,26]: ws2.cell(row=total_row, column=c).number_format = THB_FMT
    elif c in [15,27]: ws2.cell(row=total_row, column=c).number_format = PCT_FMT
    elif c == 28: ws2.cell(row=total_row, column=c).number_format = '0.00'

# ==========================================
# Sheet 4: 店铺汇总
# ==========================================
ws3 = wb.create_sheet('03-店铺汇总')
ws3.column_dimensions['A'].width = 4
ws3.column_dimensions['B'].width = 20

merge_title(ws3, 1, 1, 15, '店铺利润汇总对比表')
ws3.row_dimensions[1].height = 30

shop_headers = [
    (2, '店铺名称', 20),
    (3, '订单量', 10),
    (4, 'GMV', 14),
    (5, '有效销售额', 14),
    (6, '净销售收入', 14),
    (7, '商品成本', 14),
    (8, '平台费用', 14),
    (9, '推广费用', 14),
    (10, '达人佣金', 12),
    (11, '物流费用', 12),
    (12, '售后损失', 12),
    (13, '经营贡献利润', 16),
    (14, '经营利润率', 12),
    (15, '店铺销售占比', 12),
]

row = 2
for col, name, width in shop_headers:
    set_cell(ws3, row, col, name, header_font, header_fill, center_align, thin_border)
    ws3.column_dimensions[get_column_letter(col)].width = width

# 示例店铺行（空模板，等数据填入后可用数据透视或公式）
for r in range(3, 13):
    for c in range(2, 16):
        set_cell(ws3, r, c, '', normal_font, None, center_align if c in [3,14,15] else right_align, thin_border)
        if c in [4,5,6,7,8,9,10,11,12,13]: ws3.cell(row=r, column=c).number_format = THB_FMT
        if c in [14,15]: ws3.cell(row=r, column=c).number_format = PCT_FMT

# 合计行
set_cell(ws3, 13, 2, '合计', total_font, total_fill, center_align, thin_border)
for c in range(3, 14):
    col_letter = get_column_letter(c)
    set_cell(ws3, 13, c, f'=SUM({col_letter}3:{col_letter}12)', total_font, total_fill, right_align, thin_border)
    if c in [4,5,6,7,8,9,10,11,12,13]: ws3.cell(row=13, column=c).number_format = THB_FMT
    if c == 3: ws3.cell(row=13, column=c).number_format = INT_FMT
set_cell(ws3, 13, 14, '=IFERROR(M13/E13,0)', total_font, total_fill, center_align, thin_border)
ws3.cell(row=13, column=14).number_format = PCT_FMT
set_cell(ws3, 13, 15, '100%', total_font, total_fill, center_align, thin_border)

ws3.freeze_panes = 'C3'

# ==========================================
# Sheet 5: 成本结构分析
# ==========================================
ws4 = wb.create_sheet('04-成本结构')
ws4.column_dimensions['A'].width = 4
ws4.column_dimensions['B'].width = 24
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 40

merge_title(ws4, 1, 1, 5, '成本结构与费用占比分析')
ws4.row_dimensions[1].height = 30

cost_analysis_headers = ['成本项目', '金额(THB)', '占有效销售额比', '说明']
row = 2
for i, h in enumerate(cost_analysis_headers):
    set_cell(ws4, row, i+2, h, header_font, header_fill, center_align, thin_border)

row = 3
cost_detail = [
    ('商品成本', '=SUM(利润明细!M:M)', '采购成本/出厂价'),
    ('平台费用', '=SUM(利润明细!P:P)+SUM(利润明细!Q:Q)+SUM(利润明细!R:R)', '佣金+交易手续费+税费'),
    ('推广费用', '=SUM(利润明细!S:S)', 'TikTok Ads + GMV Max'),
    ('达人佣金', '=SUM(利润明细!T:T)', '达人/Affiliate佣金'),
    ('物流费用', '=SUM(利润明细!U:U)+SUM(利润明细!V:V)', '头程+尾程'),
    ('仓储费用', '=SUM(利润明细!W:W)', '海外仓仓储'),
    ('售后损失', '=SUM(利润明细!Y:Y)', '退款中商家承担部分'),
    ('其他经营费用', '=SUM(利润明细!X:X)', '包材/中转/提现费等'),
]
for name, formula, note in cost_detail:
    set_cell(ws4, row, 2, name, normal_font, None, left_align, thin_border)
    set_cell(ws4, row, 3, formula, normal_font, None, right_align, thin_border)
    ws4.cell(row=row, column=3).number_format = THB_FMT
    set_cell(ws4, row, 4, f'=IFERROR(C{row}/SUM(利润明细!H:H),0)', normal_font, None, center_align, thin_border)
    ws4.cell(row=row, column=4).number_format = PCT_FMT
    set_cell(ws4, row, 5, note, note_font, None, left_align, thin_border)
    row += 1

# 总变动成本
set_cell(ws4, row, 2, '总变动成本', total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), left_align, thin_border)
set_cell(ws4, row, 3, '=SUM(C3:C10)', total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), right_align, thin_border)
ws4.cell(row=row, column=3).number_format = THB_FMT
set_cell(ws4, row, 4, f'=IFERROR(C{row}/SUM(利润明细!H:H),0)', total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), center_align, thin_border)
ws4.cell(row=row, column=4).number_format = PCT_FMT
set_cell(ws4, row, 5, '', total_font, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'), left_align, thin_border)
row += 2

# 二级费用拆解
merge_section(ws4, row, 2, 5, '🔍 平台费用明细')
row += 1
platform_fees = [
    ('平台佣金', '=SUM(利润明细!P:P)', '按类目/站点不同'),
    ('交易手续费', '=SUM(利润明细!Q:Q)', '支付/交易通道费'),
    ('平台税费(VAT)', '=SUM(利润明细!R:R)', '泰国7%增值税'),
]
for name, formula, note in platform_fees:
    set_cell(ws4, row, 2, '    ' + name, normal_font, None, left_align, thin_border)
    set_cell(ws4, row, 3, formula, normal_font, None, right_align, thin_border)
    ws4.cell(row=row, column=3).number_format = THB_FMT
    set_cell(ws4, row, 4, f'=IFERROR(C{row}/SUM(利润明细!H:H),0)', normal_font, None, center_align, thin_border)
    ws4.cell(row=row, column=4).number_format = PCT_FMT
    set_cell(ws4, row, 5, note, note_font, None, left_align, thin_border)
    row += 1

row += 1
merge_section(ws4, row, 2, 5, '🚚 物流费用明细')
row += 1
logistics_fees = [
    ('头程物流', '=SUM(利润明细!U:U)', '国内→泰国运输'),
    ('尾程物流', '=SUM(利润明细!V:V)', '泰国本地配送'),
]
for name, formula, note in logistics_fees:
    set_cell(ws4, row, 2, '    ' + name, normal_font, None, left_align, thin_border)
    set_cell(ws4, row, 3, formula, normal_font, None, right_align, thin_border)
    ws4.cell(row=row, column=3).number_format = THB_FMT
    set_cell(ws4, row, 4, f'=IFERROR(C{row}/SUM(利润明细!H:H),0)', normal_font, None, center_align, thin_border)
    ws4.cell(row=row, column=4).number_format = PCT_FMT
    set_cell(ws4, row, 5, note, note_font, None, left_align, thin_border)
    row += 1

# ==========================================
# Sheet 6: 推广分析
# ==========================================
ws5 = wb.create_sheet('05-推广分析')
ws5.column_dimensions['A'].width = 4
ws5.column_dimensions['B'].width = 22
ws5.column_dimensions['C'].width = 16
ws5.column_dimensions['D'].width = 16
ws5.column_dimensions['E'].width = 16
ws5.column_dimensions['F'].width = 16
ws5.column_dimensions['G'].width = 16
ws5.column_dimensions['H'].width = 16

merge_title(ws5, 1, 1, 8, '推广效果与利润分析')
ws5.row_dimensions[1].height = 30

merge_section(ws5, 2, 2, 8, '📊 推广总览')
ad_headers = ['指标', 'TikTok Ads', 'GMV Max', '达人推广', '合计', '说明']
row = 3
for i, h in enumerate(ad_headers):
    set_cell(ws5, row, i+2, h, header_font, header_fill, center_align, thin_border)

row = 4
ad_metrics = [
    ('广告消耗', '', '', '', '', ''),
    ('广告带来销售额', '', '', '', '', ''),
    ('ROI(销售额/消耗)', '', '', '', '', '销售额÷广告消耗'),
    ('推广费用率', '', '', '', '', '广告消耗÷有效销售额'),
    ('单均推广成本', '', '', '', '', '广告消耗÷订单量'),
    ('推广带来的毛利', '', '', '', '', '需结合商品成本'),
    ('广告后真实利润', '', '', '', '', '扣除所有成本后的净利'),
]
for name, *rest in ad_metrics:
    set_cell(ws5, row, 2, name, normal_font, None, left_align, thin_border)
    for i in range(5):
        set_cell(ws5, row, i+3, '', normal_font, None, right_align, thin_border)
    set_cell(ws5, row, 8, rest[-1] if rest else '', note_font, None, left_align, thin_border)
    row += 1

row += 1
merge_section(ws5, row, 2, 8, '⚠️ 推广风险提示（自动判断）')
row += 1
risk_items = [
    ('高消耗低ROI SKU', '推广费用高但产出低的SKU清单', 'ROI < 1.5 预警'),
    ('推广费用率异常升高', '推广占比超过合理范围', '推广费用率 > 30% 预警'),
    ('ROI下降趋势', '环比下降明显', '需结合历史数据'),
    ('高销量低利润SKU', '广告驱动的销量但实际亏损', '需结合利润表'),
]
for name, desc, rule in risk_items:
    set_cell(ws5, row, 2, name, normal_font, None, left_align, thin_border)
    ws5.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    set_cell(ws5, row, 3, desc, normal_font, None, left_align, thin_border)
    ws5.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    set_cell(ws5, row, 7, rule, note_font, None, left_align, thin_border)
    row += 1

# ==========================================
# Sheet 7: 达人分析
# ==========================================
ws6 = wb.create_sheet('06-达人分析')
ws6.column_dimensions['A'].width = 4
ws6.column_dimensions['B'].width = 20
ws6.column_dimensions['C'].width = 14
ws6.column_dimensions['D'].width = 14
ws6.column_dimensions['E'].width = 14
ws6.column_dimensions['F'].width = 14
ws6.column_dimensions['G'].width = 14
ws6.column_dimensions['H'].width = 14

merge_title(ws6, 1, 1, 8, '达人渠道分析')
ws6.row_dimensions[1].height = 30

merge_section(ws6, 2, 2, 8, '📈 达人渠道总览')
row = 3
aff_headers = ['指标', '金额/数值', '占比', '说明']
for i, h in enumerate(aff_headers):
    set_cell(ws6, row, i+2, h, header_font, header_fill, center_align, thin_border)

aff_metrics = [
    ('达人渠道GMV', '', '', '达人带来的成交总额'),
    ('达人渠道有效销售额', '', '', '扣除退款后'),
    ('达人佣金总额', '', '', '实际支付的佣金'),
    ('达人佣金率', '', '', '佣金÷达人渠道有效销售额'),
    ('达人渠道商品成本', '', '', ''),
    ('达人渠道毛利', '', '', ''),
    ('达人渠道经营利润', '', '', '扣除所有成本'),
    ('达人渠道利润率', '', '', '÷ 达人渠道有效销售额'),
]
row = 4
for name, v1, v2, note in aff_metrics:
    set_cell(ws6, row, 2, name, normal_font, None, left_align, thin_border)
    set_cell(ws6, row, 3, v1, normal_font, None, right_align, thin_border)
    set_cell(ws6, row, 4, v2, normal_font, None, center_align, thin_border)
    ws6.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    set_cell(ws6, row, 5, note, note_font, None, left_align, thin_border)
    row += 1

# ==========================================
# Sheet 8: 售后分析
# ==========================================
ws7 = wb.create_sheet('07-售后分析')
ws7.column_dimensions['A'].width = 4
ws7.column_dimensions['B'].width = 24
ws7.column_dimensions['C'].width = 16
ws7.column_dimensions['D'].width = 16
ws7.column_dimensions['E'].width = 40

merge_title(ws7, 1, 1, 5, '售后损失分析')
ws7.row_dimensions[1].height = 30

aftersale_headers = ['项目', '金额/数量', '比率', '说明']
row = 2
for i, h in enumerate(aftersale_headers):
    set_cell(ws7, row, i+2, h, header_font, header_fill, center_align, thin_border)

row = 3
aftersale_items = [
    ('退款总额', '', '', '所有退款金额'),
    ('退货退款金额', '', '', '商品退回的退款'),
    ('仅退款金额', '', '', '未退货直接退款'),
    ('取消订单金额', '', '', '发货前取消'),
    ('售后补偿/赔付', '', '', '平台仲裁/补偿'),
    ('商家承担售后损失', '=SUM(利润明细!Y:Y)', '', '商家实际承担的净损失'),
    ('售后率(按金额)', '', '退款÷GMV', ''),
    ('售后损失率', '', '售后损失÷有效销售额', ''),
    ('退款订单数', '', '', ''),
    ('售后率(按订单)', '', '退款订单÷总订单', ''),
]
for item in aftersale_items:
    name = item[0]
    val = item[1]
    ratio = item[2] if len(item) > 2 else ''
    note = item[3] if len(item) > 3 else ''
    set_cell(ws7, row, 2, name, normal_font, None, left_align, thin_border)
    set_cell(ws7, row, 3, val, normal_font, None, right_align, thin_border)
    if val and val.startswith('='): ws7.cell(row=row, column=3).number_format = THB_FMT
    set_cell(ws7, row, 4, ratio, normal_font, None, center_align, thin_border)
    if '%' in str(ratio): ws7.cell(row=row, column=4).number_format = PCT_FMT
    set_cell(ws7, row, 5, note, note_font, None, left_align, thin_border)
    row += 1

# ==========================================
# Sheet 9: 指标定义
# ==========================================
ws8 = wb.create_sheet('08-指标定义')
ws8.column_dimensions['A'].width = 4
ws8.column_dimensions['B'].width = 22
ws8.column_dimensions['C'].width = 50
ws8.column_dimensions['D'].width = 20

merge_title(ws8, 1, 1, 4, '指标定义与计算公式')
ws8.row_dimensions[1].height = 30

def_headers = ['指标名称', '计算公式 / 定义', '数据来源']
row = 2
for i, h in enumerate(def_headers):
    set_cell(ws8, row, i+2, h, header_font, header_fill, center_align, thin_border)

definitions = [
    ('GMV', '所有订单的成交总额，包含退款、取消', 'TikTok销售报表'),
    ('有效销售额', 'GMV - 退款金额 - 取消金额', 'TikTok销售报表'),
    ('净销售收入', '有效销售额 - 卖家承担的折扣/促销', 'TikTok销售报表'),
    ('平台结算收入', '平台实际打款到账的金额', 'TikTok结算报表'),
    ('商品成本', '采购成本/出厂价 × 销量', 'SKU成本表'),
    ('毛利润', '净销售收入 - 商品成本', '计算得出'),
    ('毛利率', '毛利润 ÷ 有效销售额 × 100%', '计算得出'),
    ('平台佣金', '按类目比例收取的平台服务费', 'TikTok结算报表'),
    ('交易手续费', '支付/交易通道费用', 'TikTok结算报表'),
    ('平台税费', 'VAT增值税等平台代扣税费', 'TikTok结算报表'),
    ('推广费用', 'TikTok Ads + GMV Max + 其他推广', 'TikTok广告后台'),
    ('达人佣金', '达人/Affiliate带货佣金', 'TikTok联盟报表'),
    ('头程物流', '国内发往泰国的运输费用', '物流账单'),
    ('尾程物流', '泰国本地派送费用', 'TikTok结算/物流报表'),
    ('仓储费用', '海外仓仓储、操作费等', '仓储账单'),
    ('售后损失', '退款退货中商家承担的净损失', 'TikTok售后报表'),
    ('总变动成本', '商品+平台+推广+达人+物流+仓储+售后+其他', '计算得出'),
    ('经营贡献利润', '净销售收入 - 总变动成本', '计算得出'),
    ('经营贡献利润率', '经营贡献利润 ÷ 有效销售额 × 100%', '计算得出'),
    ('ROI', '有效销售额 ÷ (商品成本 + 推广费用)', '计算得出'),
    ('广告ROI', '广告带来的销售额 ÷ 广告消耗', 'TikTok广告后台'),
    ('推广费用率', '推广费用 ÷ 有效销售额 × 100%', '计算得出'),
    ('售后率(金额)', '退款金额 ÷ GMV × 100%', '计算得出'),
    ('售后损失率', '售后损失 ÷ 有效销售额 × 100%', '计算得出'),
    ('客单价', '有效销售额 ÷ 订单量', '计算得出'),
]
row = 3
for name, formula, source in definitions:
    set_cell(ws8, row, 2, name, normal_font, None, left_align, thin_border)
    set_cell(ws8, row, 3, formula, normal_font, None, left_align, thin_border)
    set_cell(ws8, row, 4, source, note_font, None, center_align, thin_border)
    row += 1

# 口径说明
row += 1
merge_section(ws8, row, 2, 4, '📌 重要口径说明')
row += 1
notes_list = [
    '1. 所有"率"类指标，分母统一使用「有效销售额」，避免口径混乱',
    '2. GMV只是流水指标，不用于利润率计算',
    '3. 平台费用以TikTok实际扣款为准，不用理论费率推算',
    '4. 汇率统一口径：建议使用账期内平均汇率，或全部按结算日实际到账汇率',
    '5. 未结算订单数据仅供预估，正式利润以已结算订单为准',
    '6. 固定成本（人工/房租/软件等）暂不计入经营贡献利润',
    '7. 同一指标不同来源数据不一致时，以结算报表为准',
]
for note in notes_list:
    ws8.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    set_cell(ws8, row, 2, note, normal_font, None, left_align, thin_border)
    row += 1

# 保存
output_path = '/app/data/所有对话/主对话/TH经营数据Agent/TikTok_Shop经营利润核算表_V1.0.xlsx'
wb.save(output_path)
print(f'生成成功: {output_path}')
